import json
import time
import openpyxl


def get_descriptoin(article, session):
    max_retries = 3
    retries = 0

    json_data = {
        'pn': f'f{article}',
        'fr': {
            'businessRegion': 1241,
        },
        'locale': 'ru-RU',
    }

    while retries < max_retries:
        r = session.post(
            'https://partscatalog.deere.com/jdrc-services/v1/partdetail/partinfo',
            json=json_data,
        )

        if r.status_code == 200:
            data = json.loads(r.text)
            return data['description']

        retries += 1
        time.sleep(2)

    if retries == max_retries:
        print(f"Can't get 'description': {article}")


def get_part_description(article, session):
    max_retries = 3
    retries = 0

    json_data = {
        'locale-cd': 'ru-RU',
        'iso2-country-code': 'RU',
        'include-sub-part-detail': 'true',
        'part-number': f'{article}',
    }

    while retries < max_retries:
        r = session.post(
            'https://partscatalog.deere.com/jdrc-services/v1/integration/parts',
            json=json_data,
        )

        if r.status_code == 200:
            data = json.loads(r.text)
            try:
                error_description = data['partOps'][0]['partBasicInfo']
                if error_description == None:
                    return False, None
            except:
                ...
            try:
                data = data['partOps'][0]['partBasicInfo']['partDescription']
                return True, data
            except:
                ...
            try:
                data = data['partOps'][1]['partBasicInfo']['partDescription']
                return True, data
            except:
                ...
        retries += 1
        time.sleep(3)

    if retries == max_retries:
        print(f"Can't get 'partDescription': {article}")


def open_file_articles():
    wb = openpyxl.load_workbook('articles.xlsx')
    ws = wb.active
    article_list = [cell.value for cell in ws['A'] if cell.value]
    return article_list


def open_file_proxy():
    with open('proxy.txt', 'r') as file:
        proxy_list = [line.strip() for line in file]
    return proxy_list


def write_data(article_list, data):
    wb = openpyxl.load_workbook('articles.xlsx')
    ws = wb.active
    for article in data:
        article_row = article_list.index(article) + 1
        ws[f'B{article_row}'] = data[article][0]
        ws[f'C{article_row}'] = data[article][1]
    wb.save('articles.xlsx')
